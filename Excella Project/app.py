import os
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CONFIG_PATH = Path(__file__).parent / "config.xml"


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOGIC
# ══════════════════════════════════════════════════════════════════════════════

def load_config():
    if not CONFIG_PATH.exists():
        messagebox.showerror("Error", "config.xml missing in folder!")
        return []
    root = ET.parse(CONFIG_PATH).getroot()
    return [(c.get("match"), c.get("label")) for c in root.findall("ColumnMapping/Column")]


def find_header_row(rows, col_matches):
    for i, row in enumerate(rows[:15]):
        cells = [str(v).strip().lower() for v in row if v is not None]
        hits = sum(1 for m in col_matches if any(m.lower() in c for c in cells))
        if hits >= 2:
            return i
    return 0


def read_sheet(filepath, sheet_name, columns):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    matches = [m for m, _ in columns]
    hdr_idx = find_header_row(rows, matches)
    headers = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
    col_map = {label: next((i for i, h in enumerate(headers) if match.lower() in h.lower()), None)
               for match, label in columns}
    records = []
    for row in rows[hdr_idx + 1:]:
        if not any(row): continue
        rec = {label: str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else ""
               for label, idx in col_map.items()}
        if rec.get("Field Name") or rec.get("Field Name in extract"):
            records.append(rec)
    return records


def generate_test_script(records, sheet_name, columns, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Script"
    BLUE_DARK = PatternFill("solid", fgColor="1F4E79")
    BLUE_MED = PatternFill("solid", fgColor="2E75B6")
    YELLOW = PatternFill("solid", fgColor="FFFACD")
    WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF")

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15

    curr = 1
    for rec in records:
        f_name = rec.get('Field Name') or rec.get('Field Name in extract')
        ws.merge_cells(f"A{curr}:E{curr}")
        cell = ws[f"A{curr}"]
        cell.value = f" Field: {f_name} | DE Caption: {rec.get('DE Caption')}"
        cell.fill, cell.font = BLUE_MED, WHITE_FONT
        curr += 1
        headers = ["TC#", "Test Case", "Input Value", "Actual Result", "Pass/Fail"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=curr, column=i, value=h)
            c.fill, c.font, c.alignment = BLUE_DARK, WHITE_FONT, Alignment(horizontal="center")
        curr += 1
        test_types = [lbl for _, lbl in columns if any(x in lbl for x in ["Positive", "Negative"])]
        for i, ttype in enumerate(test_types, 1):
            ws.cell(row=curr, column=1, value=i).alignment = Alignment(horizontal="center")
            ws.cell(row=curr, column=2, value=ttype)
            ws.cell(row=curr, column=3, value=rec.get(ttype))
            ws.cell(row=curr, column=4).fill = YELLOW
            ws.cell(row=curr, column=5).fill = YELLOW
            curr += 1
        curr += 1
    wb.save(output_path)


# ══════════════════════════════════════════════════════════════════════════════
# UI - MATCHING IMAGE FORMAT
# ══════════════════════════════════════════════════════════════════════════════

class App:
    def __init__(self, root):
        self.root = root
        self.columns = load_config()
        self.filepath = ""

        root.title("XBP - Excel Test Script Generator")
        root.geometry("700x350")
        root.configure(bg="white")
        root.resizable(False, False)

        self._build_ui()

    def _build_ui(self):
        # Top Logo & Title Section
        header_frame = tk.Frame(self.root, bg="white")
        header_frame.pack(fill="x", pady=20, padx=30)

        # Logo (Simulated XBP label)
        logo_label = tk.Label(header_frame, text="⇌XBP", font=("Arial", 28, "bold"),
                              bg="black", fg="white", padx=10)
        logo_label.pack(side="left")

        title_label = tk.Label(header_frame, text="Excel Test Script Generator",
                               font=("Arial", 22, "bold"), bg="white", fg="black")
        title_label.pack(side="left", padx=20)

        # Body Section
        body = tk.Frame(self.root, bg="white", padx=50)
        body.pack(fill="both", expand=True)

        # File Path Row
        tk.Label(body, text="Excel File Path:", font=("Arial", 10, "bold"),
                 bg="white", fg="#006400").grid(row=0, column=0, sticky="e", pady=15)

        self.path_var = tk.StringVar(value="")
        self.entry_path = tk.Entry(body, textvariable=self.path_var, width=45,
                                   bg="#eeeeee", relief="solid", bd=1)
        self.entry_path.grid(row=0, column=1, padx=10)

        tk.Button(body, text="...", command=self.browse, width=3, relief="solid", bd=1).grid(row=0, column=2)

        # Sheet Selection Row
        tk.Label(body, text="Select Excel Sheet:", font=("Arial", 10, "bold"),
                 bg="white", fg="#006400").grid(row=1, column=0, sticky="e", pady=15)

        self.sheet_var = tk.StringVar()
        self.sheet_cb = ttk.Combobox(body, textvariable=self.sheet_var, width=43, state="disabled")
        self.sheet_cb.grid(row=1, column=1, padx=10, sticky="w")

        # Footer Buttons
        btn_frame = tk.Frame(self.root, bg="white", pady=30)
        btn_frame.pack(fill="x")

        self.btn_create = tk.Button(btn_frame, text="Create", font=("Arial", 10, "bold"),
                                    bg="#dddddd", fg="blue", width=15, height=2,
                                    relief="raised", command=self.run, state="disabled")
        self.btn_create.pack(side="left", padx=(180, 20))

        tk.Button(btn_frame, text="Exit", font=("Arial", 10, "bold"),
                  bg="#dddddd", fg="black", width=15, height=2,
                  relief="raised", command=self.root.destroy).pack(side="left")

    def browse(self):
        file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file:
            self.filepath = file
            self.path_var.set(file)
            wb = openpyxl.load_workbook(file, read_only=True)
            self.sheet_cb.config(values=wb.sheetnames, state="readonly")
            self.sheet_cb.current(0)
            self.btn_create.config(state="normal")

    def run(self):
        try:
            records = read_sheet(self.filepath, self.sheet_cb.get(), self.columns)
            out = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                               initialfile=f"{self.sheet_cb.get()}_Script.xlsx")
            if out:
                generate_test_script(records, self.sheet_cb.get(), self.columns, out)
                messagebox.showinfo("Success", "Test script generated successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()





